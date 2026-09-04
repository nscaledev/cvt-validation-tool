"""Load Node_Table / ADMIN maps and mutate CABLING_REMEDIATION."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from cvt_sharepoint.triage import (
    DerivedRow,
    _s,
    connection_key,
    issue_key,
    reopen_note,
)

SHEET_REMEDIATION = "CABLING_REMEDIATION"
SHEET_NODE = "Node_Table"
SHEET_ADMIN = "ADMIN"

# CABLING_REMEDIATION columns (1-based)
COL_ISSUE = 1
COL_PROTOCOL = 2
COL_REPORTER = 3
COL_DATE = 4
COL_SU = 5
COL_SRC_MSFT = 6
COL_SRC_NSCALE = 7
COL_SRC_RACK = 8
COL_SRC_U = 9
COL_SRC_PORT = 10
COL_SRC_SN = 11
COL_DST_MSFT = 12
COL_DST_NSCALE = 13
COL_DST_RACK = 14
COL_DST_U = 15
COL_DST_PORT = 16
COL_DST_SN = 17
COL_ISSUE_TYPE = 18
COL_COMMENTS = 19
COL_PRIORITY = 20
COL_STATUS = 21
COL_REM_ACTION = 22
COL_REM_DATE = 23
COL_COMMENT_DELL = 24
COL_REOPEN_COUNT = 25
COL_REOPEN_NOTES = 26
COL_CONN_KEY = 27
COL_ISSUE_KEY = 28

DATA_START_ROW = 3


@dataclass
class RemediationRow:
    excel_row: int
    issue_no: Any
    protocol: str
    status: str
    src_host: str
    src_port: str
    dst_host: str
    dst_port: str
    issue_type: str
    reopen_count: Any
    reopen_notes: str
    conn_key: str
    iss_key: str
    # Columns A–Z (Issue # … Re-open Notes) for full-row SharePoint paste.
    values: list[Any] = field(default_factory=list)


@dataclass
class PrepareResult:
    tracker_path: Path
    backup_path: Path
    reopens: list[dict[str, Any]] = field(default_factory=list)
    adds_new_connection: list[dict[str, Any]] = field(default_factory=list)
    adds_new_issue: list[dict[str, Any]] = field(default_factory=list)
    skipped_tracked: int = 0
    skipped_other: int = 0
    rows_before: int = 0
    rows_after: int = 0
    next_issue_no: int = 1


def load_node_table(ws: Worksheet) -> dict[str, tuple[str, str, str]]:
    """Node_Table A=MSFT host, B=Nscale, C=Rack, D=U."""
    table: dict[str, tuple[str, str, str]] = {}
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        host = _s(row[0] if row else "")
        if not host:
            continue
        nscale = _s(row[1] if len(row) > 1 else "")
        rack = _s(row[2] if len(row) > 2 else "")
        unit = _s(row[3] if len(row) > 3 else "")
        table[host] = (nscale, rack, unit)
    return table


def load_admin_port_map(ws: Worksheet) -> dict[str, str]:
    """ADMIN O:P port remap used by CVT_IMPORT formulas."""
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=3, max_row=50, min_col=15, max_col=16, values_only=True):
        src = _s(row[0] if row else "")
        dst = _s(row[1] if row and len(row) > 1 else "")
        if src:
            mapping[src] = dst or src
    return mapping


def _build_keys_for_row(
    src_host: str,
    src_port: str,
    dst_host: str,
    dst_port: str,
    issue_type: str,
    conn_key_cell: Any,
    issue_key_cell: Any,
) -> tuple[str, str]:
    conn = _s(conn_key_cell)
    iss = _s(issue_key_cell)
    if not conn:
        conn = connection_key(src_host, src_port, dst_host, dst_port)
    if not iss:
        iss = issue_key(conn, issue_type)
    return conn, iss


def load_remediation_index(
    ws: Worksheet,
    *,
    on_progress: Callable[[int, int], None] | None = None,
) -> tuple[list[RemediationRow], dict[str, str], set[str], dict[str, RemediationRow], int, dict[str, str]]:
    from collections import Counter

    from cvt_sharepoint.comments import clean_engineer_comment, finalize_nvidia_hints

    rows: list[RemediationRow] = []
    issue_key_to_status: dict[str, str] = {}
    issue_key_to_row: dict[str, RemediationRow] = {}
    max_issue = 0
    hint_tallies: dict[str, Counter[str]] = {}

    approx_total = max(1, (ws.max_row or 5000) - DATA_START_ROW + 1)
    scanned = 0

    for excel_row, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, max_col=COL_ISSUE_KEY, values_only=True),
        start=DATA_START_ROW,
    ):
        scanned += 1
        if on_progress and (scanned == 1 or scanned % 200 == 0):
            on_progress(min(scanned, approx_total), approx_total)
        issue_no = row[COL_ISSUE - 1] if row else None
        src_host = _s(row[COL_SRC_MSFT - 1] if row and len(row) >= COL_SRC_MSFT else "")
        if issue_no is None and not src_host:
            continue
        src_port = _s(row[COL_SRC_PORT - 1])
        dst_host = _s(row[COL_DST_MSFT - 1])
        dst_port = _s(row[COL_DST_PORT - 1])
        issue_type = _s(row[COL_ISSUE_TYPE - 1])
        status = _s(row[COL_STATUS - 1])
        conn_cell = row[COL_CONN_KEY - 1] if len(row) >= COL_CONN_KEY else None
        iss_cell = row[COL_ISSUE_KEY - 1] if len(row) >= COL_ISSUE_KEY else None
        conn, iss = _build_keys_for_row(
            src_host,
            src_port,
            dst_host,
            dst_port,
            issue_type,
            conn_cell,
            iss_cell,
        )
        # Full A–Z row for handoff paste (pad to Re-open Notes).
        raw_vals = list(row[:COL_REOPEN_NOTES]) if row else []
        while len(raw_vals) < COL_REOPEN_NOTES:
            raw_vals.append(None)
        rec = RemediationRow(
            excel_row=excel_row,
            issue_no=issue_no,
            protocol=_s(row[COL_PROTOCOL - 1]),
            status=status,
            src_host=src_host,
            src_port=src_port,
            dst_host=dst_host,
            dst_port=dst_port,
            issue_type=issue_type,
            reopen_count=row[COL_REOPEN_COUNT - 1],
            reopen_notes=_s(row[COL_REOPEN_NOTES - 1]),
            conn_key=conn,
            iss_key=iss,
            values=raw_vals,
        )
        rows.append(rec)
        if iss:
            issue_key_to_status[iss] = status
            issue_key_to_row[iss] = rec
        if issue_type:
            comment = clean_engineer_comment(row[COL_COMMENTS - 1] if len(row) >= COL_COMMENTS else "")
            if comment:
                hint_tallies.setdefault(issue_type, Counter())[comment] += 1
        try:
            max_issue = max(max_issue, int(issue_no))
        except (TypeError, ValueError):
            pass

    if on_progress:
        on_progress(approx_total, approx_total)

    conn_keys = {r.conn_key for r in rows if r.conn_key}
    nvidia_hints = finalize_nvidia_hints(hint_tallies)
    return rows, issue_key_to_status, conn_keys, issue_key_to_row, max_issue, nvidia_hints


def _inc_reopen_count(value: Any) -> int:
    try:
        return int(value) + 1
    except (TypeError, ValueError):
        return 1


def apply_reopen(ws: Worksheet, rec: RemediationRow, when: datetime | None = None) -> dict[str, Any]:
    from cvt_sharepoint.triage import reported_at, reporter_initials

    when = when or datetime.now()
    old_status = rec.status
    old_count = rec.reopen_count
    old_notes = rec.reopen_notes
    new_count = _inc_reopen_count(old_count)
    initials = reporter_initials()
    new_notes = reopen_note(
        old_notes,
        initials=initials,
        issue_type=rec.issue_type,
        when=when,
    )
    new_status = "Re-opened"
    ws.cell(rec.excel_row, COL_REPORTER, initials)
    ws.cell(rec.excel_row, COL_DATE, reported_at(when))
    ws.cell(rec.excel_row, COL_STATUS, new_status)
    ws.cell(rec.excel_row, COL_REOPEN_COUNT, new_count)
    ws.cell(rec.excel_row, COL_REOPEN_NOTES, new_notes)
    return {
        "issue_no": rec.issue_no,
        "excel_row": rec.excel_row,
        "conn_key": rec.conn_key,
        "issue_type": rec.issue_type,
        "old_status": old_status,
        "new_status": new_status,
        "old_count": old_count,
        "new_count": new_count,
        "old_notes": old_notes,
        "new_notes": new_notes,
        "src_host": rec.src_host,
        "src_port": rec.src_port,
        "dst_host": rec.dst_host,
        "dst_port": rec.dst_port,
    }


def append_add_row(ws: Worksheet, excel_row: int, issue_no: int, derived: DerivedRow) -> dict[str, Any]:
    values = {
        COL_ISSUE: issue_no,
        COL_PROTOCOL: derived.protocol,
        COL_REPORTER: None,
        COL_DATE: None,
        COL_SU: derived.su,
        COL_SRC_MSFT: derived.src_host,
        COL_SRC_NSCALE: derived.src_nscale or None,
        COL_SRC_RACK: derived.src_rack or None,
        COL_SRC_U: derived.src_u or None,
        COL_SRC_PORT: derived.src_port,
        COL_SRC_SN: None,
        COL_DST_MSFT: derived.dst_host,
        COL_DST_NSCALE: derived.dst_nscale or None,
        COL_DST_RACK: derived.dst_rack or None,
        COL_DST_U: derived.dst_u or None,
        COL_DST_PORT: derived.dst_port,
        COL_DST_SN: None,
        COL_ISSUE_TYPE: derived.issue_type,
        COL_COMMENTS: derived.comments_nvidia or None,
        COL_PRIORITY: None,
        COL_STATUS: derived.status_default,
        COL_REM_ACTION: None,
        COL_REM_DATE: None,
        COL_COMMENT_DELL: None,
        COL_REOPEN_COUNT: None,
        COL_REOPEN_NOTES: None,
        # Leave AA/AB formula spill alone — write keys as values only if spill broken;
        # writing values into spill area can break array formulas. Skip AA/AB.
    }
    for col, val in values.items():
        ws.cell(excel_row, col, val)
    return {
        "issue_no": issue_no,
        "excel_row": excel_row,
        "triage": derived.triage,
        "protocol": derived.protocol,
        "su": derived.su,
        "src_host": derived.src_host,
        "src_port": derived.src_port,
        "dst_host": derived.dst_host,
        "dst_port": derived.dst_port,
        "issue_type": derived.issue_type,
        "status": derived.status_default,
        "conn_key": derived.conn_key,
    }


def first_empty_data_row(ws: Worksheet) -> int:
    max_row = ws.max_row or DATA_START_ROW
    for excel_row in range(DATA_START_ROW, max_row + 2):
        if ws.cell(excel_row, COL_ISSUE).value is None and not _s(ws.cell(excel_row, COL_SRC_MSFT).value):
            return excel_row
    return max_row + 1


def process_workbook(
    tracker_path: Path,
    derived_rows: list[DerivedRow],
    *,
    dry_run: bool = False,
) -> PrepareResult:
    wb = load_workbook(tracker_path)
    if SHEET_REMEDIATION not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet {SHEET_REMEDIATION}")
    rem = wb[SHEET_REMEDIATION]
    node_table = load_node_table(wb[SHEET_NODE]) if SHEET_NODE in wb.sheetnames else {}
    admin_map = load_admin_port_map(wb[SHEET_ADMIN]) if SHEET_ADMIN in wb.sheetnames else {}

    # Re-derive with node/admin from this workbook if callers passed incomplete lookups
    _ = node_table, admin_map

    rows, issue_status, conn_keys, issue_to_row, max_issue, _hints = load_remediation_index(rem)
    result = PrepareResult(
        tracker_path=tracker_path,
        backup_path=tracker_path,  # filled by caller
        rows_before=len(rows),
        next_issue_no=max_issue + 1,
    )

    next_issue = max_issue + 1
    append_at = first_empty_data_row(rem)
    seen_reopen: set[str] = set()
    seen_add: set[str] = set()

    for derived in derived_rows:
        if derived.triage == "REOPEN EXISTING":
            if derived.iss_key in seen_reopen:
                continue
            seen_reopen.add(derived.iss_key)
            rec = issue_to_row.get(derived.iss_key)
            if rec is None:
                result.skipped_other += 1
                continue
            if dry_run:
                result.reopens.append(
                    {
                        "issue_no": rec.issue_no,
                        "excel_row": rec.excel_row,
                        "conn_key": rec.conn_key,
                        "issue_type": rec.issue_type,
                        "old_status": rec.status,
                        "new_status": "Re-opened",
                        "old_count": rec.reopen_count,
                        "new_count": _inc_reopen_count(rec.reopen_count),
                        "src_host": rec.src_host,
                        "src_port": rec.src_port,
                        "dst_host": rec.dst_host,
                        "dst_port": rec.dst_port,
                    }
                )
            else:
                result.reopens.append(apply_reopen(rem, rec))
        elif derived.triage in {"ADD - NEW CONNECTION", "ADD - NEW ISSUE TYPE"}:
            if derived.iss_key in seen_add:
                continue
            # Also skip if somehow already present
            if derived.iss_key in issue_to_row:
                result.skipped_tracked += 1
                continue
            seen_add.add(derived.iss_key)
            info = {
                "issue_no": next_issue,
                "excel_row": append_at,
                "triage": derived.triage,
                "protocol": derived.protocol,
                "su": derived.su,
                "src_host": derived.src_host,
                "src_port": derived.src_port,
                "dst_host": derived.dst_host,
                "dst_port": derived.dst_port,
                "issue_type": derived.issue_type,
                "status": derived.status_default,
                "conn_key": derived.conn_key,
            }
            if not dry_run:
                info = append_add_row(rem, append_at, next_issue, derived)
            if derived.triage == "ADD - NEW CONNECTION":
                result.adds_new_connection.append(info)
            else:
                result.adds_new_issue.append(info)
            next_issue += 1
            append_at += 1
        elif derived.triage == "ALREADY TRACKED":
            result.skipped_tracked += 1
        else:
            result.skipped_other += 1

    result.rows_after = result.rows_before + len(result.adds_new_connection) + len(result.adds_new_issue)
    result.next_issue_no = next_issue
    if not dry_run:
        wb.save(tracker_path)
    wb.close()
    return result
