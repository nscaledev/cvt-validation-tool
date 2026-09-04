from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


class LocalWorkbookError(RuntimeError):
    pass


def local_workbook_path() -> Path | None:
    raw = (os.environ.get("SHAREPOINT_LOCAL_PATH") or "").strip()
    if not raw:
        return None
    return Path(raw).expanduser()


def require_local_workbook() -> Path:
    path = local_workbook_path()
    if path is None:
        raise LocalWorkbookError(
            "No downloaded workbook path is set.\n"
            "1. ./cvt sharepoint open   (browser, JumpCloud if needed)\n"
            "2. In SharePoint, Download the xlsx (not Open in Desktop / OneDrive)\n"
            "3. ./cvt sharepoint use --file ~/Downloads/the-file.xlsx"
        )
    if not path.is_file():
        raise LocalWorkbookError(
            f"Downloaded workbook not found: {path}\n"
            "Download it again from SharePoint and run: ./cvt sharepoint use --file <path>"
        )
    return path


def workbook_info(path: Path) -> dict[str, Any]:
    stat = path.stat()
    mtime = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
    return {"name": path.name, "path": str(path.resolve()), "size": stat.st_size, "lastModified": mtime}


def sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def update_range(path: Path, worksheet: str, address: str, values: list[list[Any]]) -> str:
    wb = load_workbook(path)
    if worksheet not in wb.sheetnames:
        names = ", ".join(wb.sheetnames)
        wb.close()
        raise LocalWorkbookError(f"Worksheet {worksheet!r} not found. Available: {names}")
    ws = wb[worksheet]
    if ":" in address:
        min_col, min_row, _, _ = range_boundaries(address)
    else:
        col_letter, row = coordinate_from_string(address)
        min_col = column_index_from_string(col_letter)
        min_row = row
    row_vals = values[0] if values else []
    for offset, value in enumerate(row_vals):
        ws.cell(min_row, min_col + offset, value)
    wb.save(path)
    wb.close()
    return address
