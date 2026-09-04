"""Build capped audit handoff batches (HTML/Excel copy-paste, no full-tab replace)."""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

from openpyxl import Workbook, load_workbook

from cvt_sharepoint.comments import comments_for_cvt_row
from cvt_sharepoint.remediation import (
    COL_COMMENT_DELL,
    COL_DATE,
    COL_DST_NSCALE,
    COL_DST_RACK,
    COL_DST_U,
    COL_REM_ACTION,
    COL_REM_DATE,
    COL_REOPEN_COUNT,
    COL_REOPEN_NOTES,
    COL_REPORTER,
    COL_SRC_NSCALE,
    COL_SRC_RACK,
    COL_SRC_U,
    COL_STATUS,
    SHEET_ADMIN,
    SHEET_NODE,
    SHEET_REMEDIATION,
    RemediationRow,
    load_admin_port_map,
    load_node_table,
    load_remediation_index,
)
from cvt_sharepoint.triage import (
    KEY_SEP,
    DerivedRow,
    derive_row,
    lookup_node,
    reopen_note,
    reported_at,
    reporter_initials,
    _s,
)

ProgressFn = Callable[[str, int, Optional[int]], None]

REOPEN_LIMIT = 20
ADD_LIMIT = 300

# Triage labels that wrongly landed in Rem Date / Dell cols on some tracker rows.
_TRIAGE_POLLUTION = {
    "ALREADY TRACKED",
    "REOPEN EXISTING",
    "ADD - NEW CONNECTION",
    "ADD - NEW ISSUE TYPE",
}

# Exact CABLING_REMEDIATION columns A–Z (paste Values from column A).
# Reopen + adds both use this set only — never Excel row / Paste at / AA+.
REMEDIATION_HEADERS = [
    "Issue #",
    "Protocol",
    "Reporter",
    "Date Reported",
    "SU",
    "Source Hostname (MSFT)",
    "Source Hostname (Nscale)",
    "Source Rack",
    "Source U",
    "Source Port",
    "Source SN",
    "Dest Hostname (MSFT)",
    "Dest Hostname (Nscale)",
    "Dest Rack",
    "Dest U",
    "Dest Port",
    "Destination SN",
    "Issue Type",
    "Comments - NVIDIA",
    "Priority",
    "Status",
    "Remediation Action Taken",
    "Remediation Date/Time",
    "Comment Dell/Cabling",
    "Re-open Count",
    "Re-open Notes",
]

# Keep aliases used elsewhere.
ADD_HEADERS = REMEDIATION_HEADERS
REOPEN_HEADERS = REMEDIATION_HEADERS  # A–Z only; guide cols live in HTML, not Excel paste sheet


def _inc_reopen_count(value: Any) -> int:
    try:
        return int(value) + 1
    except (TypeError, ValueError):
        return 1


def _tsv(cells: list[Any]) -> str:
    parts: list[str] = []
    for cell in cells:
        text = "" if cell is None else str(cell)
        text = text.replace("\t", " ").replace("\r\n", " | ").replace("\n", " | ")
        parts.append(text)
    return "\t".join(parts)


def _paste_cell(value: Any) -> Any:
    """Normalize values for SharePoint Values paste (no formula cache junk)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y %H:%M")
    return value


def _is_blankish(value: Any) -> bool:
    if value is None:
        return True
    if value == 0 or value == "0":
        return True
    return str(value).strip() == ""


def _is_triage_pollution(value: Any) -> bool:
    """True if cell holds a CVT triage label or a spilled connection/issue key."""
    text = _s(value)
    if not text:
        return False
    if text.upper() in _TRIAGE_POLLUTION:
        return True
    # Conn/issue keys use ¦ and belong in AA/AB, not Rem Date / Comment Dell.
    if KEY_SEP in text:
        return True
    return False


def _reopen_values(
    rec: RemediationRow,
    *,
    node_table: dict,
    new_status: str,
    new_count: int,
    new_notes: str,
    reporter: str,
    date_reported: str,
) -> list[Any]:
    """Full A–Z row: keep existing data; update status / count / notes / reporter / date.

    Repair Nscale/rack/U when data_only read returned 0/blank from formulas.
    Strip triage labels / key spills that polluted Rem Date / Dell on some rows.
    """
    values = list(rec.values) if rec.values else [None] * len(REMEDIATION_HEADERS)
    while len(values) < len(REMEDIATION_HEADERS):
        values.append(None)
    values = values[: len(REMEDIATION_HEADERS)]

    src_nscale, src_rack, src_u = lookup_node(node_table, rec.src_host)
    dst_nscale, dst_rack, dst_u = lookup_node(node_table, rec.dst_host)
    if _is_blankish(values[COL_SRC_NSCALE - 1]) and src_nscale:
        values[COL_SRC_NSCALE - 1] = src_nscale
    if _is_blankish(values[COL_SRC_RACK - 1]) and src_rack:
        values[COL_SRC_RACK - 1] = src_rack
    if _is_blankish(values[COL_SRC_U - 1]) and src_u:
        values[COL_SRC_U - 1] = src_u
    if _is_blankish(values[COL_DST_NSCALE - 1]) and dst_nscale:
        values[COL_DST_NSCALE - 1] = dst_nscale
    if _is_blankish(values[COL_DST_RACK - 1]) and dst_rack:
        values[COL_DST_RACK - 1] = dst_rack
    if _is_blankish(values[COL_DST_U - 1]) and dst_u:
        values[COL_DST_U - 1] = dst_u

    # Never paste formula-cache junk (0) into hostname Nscale columns.
    if values[COL_SRC_NSCALE - 1] in (0, "0"):
        values[COL_SRC_NSCALE - 1] = src_nscale or ""
    if values[COL_DST_NSCALE - 1] in (0, "0"):
        values[COL_DST_NSCALE - 1] = dst_nscale or ""

    # Some tracker rows have triage / AA-AB keys pasted into V–X — do not re-paste them.
    if _is_triage_pollution(values[COL_REM_ACTION - 1]):
        values[COL_REM_ACTION - 1] = ""
    if _is_triage_pollution(values[COL_REM_DATE - 1]):
        values[COL_REM_DATE - 1] = ""
    if _is_triage_pollution(values[COL_COMMENT_DELL - 1]):
        values[COL_COMMENT_DELL - 1] = ""

    values[COL_REPORTER - 1] = reporter
    values[COL_DATE - 1] = date_reported
    values[COL_STATUS - 1] = new_status
    values[COL_REOPEN_COUNT - 1] = new_count
    values[COL_REOPEN_NOTES - 1] = new_notes
    return [_paste_cell(v) for v in values]


@dataclass
class HandoffResult:
    tracker_path: Path
    backup_path: Path | None
    cvt_csv: Path | None
    reopens: list[dict[str, Any]] = field(default_factory=list)
    adds: list[dict[str, Any]] = field(default_factory=list)
    triage_totals: dict[str, int] = field(default_factory=dict)
    reopen_total: int = 0
    add_total: int = 0
    skipped_tracked: int = 0
    next_issue_no: int = 1


def load_context(
    tracker: Path,
    *,
    on_progress: ProgressFn | None = None,
) -> tuple[dict, dict, dict[str, str], set[str], dict[str, RemediationRow], int, dict[str, str]]:
    def emit(phase: str, current: int = 0, total: int | None = None) -> None:
        if on_progress:
            on_progress(phase, current, total)

    emit("open", 0, None)
    # openpyxl emits this during read_only sheet iteration, not only at open.
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl\..*")
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        # read_only streams the ~50MB remediation sheet instead of a multi-minute full parse.
        wb = load_workbook(tracker, read_only=True, data_only=True)
        try:
            rem = wb[SHEET_REMEDIATION]
            emit("nodes", 0, None)
            node = load_node_table(wb[SHEET_NODE]) if SHEET_NODE in wb.sheetnames else {}
            admin = load_admin_port_map(wb[SHEET_ADMIN]) if SHEET_ADMIN in wb.sheetnames else {}

            def index_progress(current: int, total: int) -> None:
                emit("index", current, total)

            emit("index", 0, max(1, (rem.max_row or 5000) - 2))
            # Single pass: keys + full A–Z values + NVIDIA hint mining (read_only sheets iterate once).
            _, issue_status, conn_keys, issue_to_row, max_issue, nvidia_hints = load_remediation_index(
                rem, on_progress=index_progress
            )
            emit("hints", 1, 1)
        finally:
            wb.close()
    return node, admin, issue_status, conn_keys, issue_to_row, max_issue + 1, nvidia_hints


def build_handoff(
    tracker: Path,
    cvt_rows: list[dict[str, str]],
    *,
    backup_path: Path | None = None,
    cvt_csv: Path | None = None,
    reopen_limit: int = REOPEN_LIMIT,
    add_limit: int = ADD_LIMIT,
    on_progress: ProgressFn | None = None,
) -> HandoffResult:
    def emit(phase: str, current: int = 0, total: int | None = None) -> None:
        if on_progress:
            on_progress(phase, current, total)

    node, admin, issue_status, conn_keys, issue_to_row, next_issue, nvidia_hints = load_context(
        tracker, on_progress=on_progress
    )

    total_cvt = max(len(cvt_rows), 1)
    derived: list[DerivedRow] = []
    for i, row in enumerate(cvt_rows, start=1):
        if i == 1 or i % 25 == 0 or i == total_cvt:
            emit("match", i, total_cvt)
        item = derive_row(
            row,
            node_table=node,
            admin_map=admin,
            existing_issue_keys=issue_status,
            existing_conn_keys=conn_keys,
        )
        if item is not None:
            item.comments_nvidia = comments_for_cvt_row(
                item.source, item.issue_type, mined=nvidia_hints, admin_map=admin
            )
            derived.append(item)

    emit("batch", 0, None)
    totals: dict[str, int] = {}
    for item in derived:
        totals[item.triage] = totals.get(item.triage, 0) + 1

    reopens_all: list[tuple[DerivedRow, RemediationRow]] = []
    seen_reopen: set[str] = set()
    for item in derived:
        if item.triage != "REOPEN EXISTING" or item.iss_key in seen_reopen:
            continue
        rec = issue_to_row.get(item.iss_key)
        if rec is None:
            continue
        seen_reopen.add(item.iss_key)
        reopens_all.append((item, rec))
    reopens_all.sort(key=lambda pair: (int(pair[1].issue_no) if str(pair[1].issue_no).isdigit() else 10**9, pair[1].excel_row))

    adds_all: list[DerivedRow] = []
    seen_add: set[str] = set()
    for item in derived:
        if item.triage not in {"ADD - NEW CONNECTION", "ADD - NEW ISSUE TYPE"}:
            continue
        if item.iss_key in seen_add or item.iss_key in issue_to_row:
            continue
        seen_add.add(item.iss_key)
        adds_all.append(item)
    adds_all.sort(key=lambda d: (d.triage, d.src_host, d.src_port, d.dst_host, d.dst_port))

    result = HandoffResult(
        tracker_path=tracker,
        backup_path=backup_path,
        cvt_csv=cvt_csv,
        triage_totals=totals,
        reopen_total=len(reopens_all),
        add_total=len(adds_all),
        skipped_tracked=totals.get("ALREADY TRACKED", 0),
        next_issue_no=next_issue,
    )

    when = datetime.now()
    initials = reporter_initials()
    date_str = reported_at(when)
    for item, rec in reopens_all[:reopen_limit]:
        new_count = _inc_reopen_count(rec.reopen_count)
        new_notes = reopen_note(
            rec.reopen_notes,
            initials=initials,
            issue_type=item.issue_type or rec.issue_type,
            a_report=item.source.get("A Report") or item.source.get("E") or "",
            z_report=item.source.get("Z Report") or item.source.get("T") or "",
            a_remediation=item.source.get("A Remediation Action") or item.source.get("F") or "",
            z_remediation=item.source.get("Z Remediation Action") or item.source.get("U") or "",
            cvt=item.source,
            admin_map=admin,
            when=when,
        )
        new_status = "Re-opened"
        values = _reopen_values(
            rec,
            node_table=node,
            new_status=new_status,
            new_count=new_count,
            new_notes=new_notes,
            reporter=initials,
            date_reported=date_str,
        )
        result.reopens.append(
            {
                "issue_no": rec.issue_no,
                "excel_row": rec.excel_row,
                "paste_at": f"A{rec.excel_row}",
                "old_status": rec.status,
                "new_status": new_status,
                "old_count": rec.reopen_count,
                "new_count": new_count,
                "old_notes": rec.reopen_notes,
                "new_notes": new_notes,
                "src_host": rec.src_host,
                "src_port": rec.src_port,
                "dst_host": rec.dst_host,
                "dst_port": rec.dst_port,
                "issue_type": rec.issue_type,
                "headers": list(REMEDIATION_HEADERS),
                "values": values,
                "tsv": _tsv(values),
            }
        )

    issue_no = next_issue
    for item in adds_all[:add_limit]:
        comments = item.comments_nvidia
        values = [
            issue_no,
            item.protocol,
            initials,
            date_str,
            item.su,
            item.src_host,
            item.src_nscale,
            item.src_rack,
            item.src_u,
            item.src_port,
            "",
            item.dst_host,
            item.dst_nscale,
            item.dst_rack,
            item.dst_u,
            item.dst_port,
            "",
            item.issue_type,
            comments,
            "",
            item.status_default,
            "",
            "",
            "",
            "",
            "",
        ]
        result.adds.append(
            {
                "triage": item.triage,
                "issue_no": issue_no,
                "protocol": item.protocol,
                "su": item.su,
                "src_host": item.src_host,
                "src_port": item.src_port,
                "dst_host": item.dst_host,
                "dst_port": item.dst_port,
                "issue_type": item.issue_type,
                "status": item.status_default,
                "comments": comments,
                "headers": list(REMEDIATION_HEADERS),
                "values": values,
                "tsv": _tsv(values),
            }
        )
        issue_no += 1

    result.next_issue_no = issue_no
    return result


def write_handoff_xlsx(path: Path, result: HandoffResult) -> Path:
    wb = Workbook()
    ws_r = wb.active
    ws_r.title = "REOPEN"
    ws_r.append(REOPEN_HEADERS)
    for row in result.reopens:
        # A–Z only — no Excel row / Paste at (those are HTML guides).
        ws_r.append(list(row["values"]))

    ws_a = wb.create_sheet("ADDS")
    ws_a.append(ADD_HEADERS)
    for row in result.adds:
        ws_a.append(row["values"])

    ws_s = wb.create_sheet("SUMMARY")
    ws_s.append(["Field", "Value"])
    ws_s.append(["Tracker", str(result.tracker_path)])
    ws_s.append(["CVT CSV", str(result.cvt_csv or "")])
    ws_s.append(["Tracker mutated", "No — read-only keys; handoff only"])
    ws_s.append(["Reopen paste range", "A–Z on existing row (Issue # … Re-open Notes); leave AA+ alone"])
    ws_s.append(["Reopen in this batch", len(result.reopens)])
    ws_s.append(["Reopen remaining after this batch", max(0, result.reopen_total - len(result.reopens))])
    ws_s.append(["Adds in this batch", len(result.adds)])
    ws_s.append(["Adds remaining after this batch", max(0, result.add_total - len(result.adds))])
    if result.reopens:
        ws_s.append([])
        ws_s.append(["Reopen Issue #", "Excel row", "Paste at"])
        for row in result.reopens:
            ws_s.append([row["issue_no"], row["excel_row"], row["paste_at"]])

    wb.save(path)
    return path
