from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable, Sequence

# Visible Circuits View columns (Displayed Columns / CSV "Displayed Data").
# Hidden menu columns are omitted. Icon-only buttons are omitted.
DISPLAYED_COLUMNS: list[str] = [
    "Status",
    "Protocol",
    "A Data Hall",
    "A SU Number",
    "A Report",
    "A Remediation Action",
    "A Location",
    "Exp A Node",
    "Exp A Port",
    "Disc A Node",
    "Disc A Port",
    "A Transceiver Connected",
    "A Port Status",
    "A Signal Stats",
    "A Raw BER",
    "A Eff BER",
    "A Symbol BER",
    "Z Data Hall",
    "Z SU Number",
    "Z Report",
    "Z Remediation Action",
    "Z Location",
    "Exp Z Node",
    "Exp Z Port",
    "Disc Z Node",
    "Disc Z Port",
    "Z Transceiver Connected",
    "Z Port Status",
    "Z Signal Stats",
    "Z Raw BER",
    "Z Eff BER",
    "Z Symbol BER",
]

NA = "NA"


def _na(value: Any) -> Any:
    if value is None or value == "":
        return NA
    return value


def _null_as_na(value: Any) -> str:
    if value is None:
        return NA
    return str(value)


def _endpoint_port(endpoint: dict[str, Any] | None) -> Any:
    endpoint = endpoint or {}
    return endpoint.get("display_port") or endpoint.get("port")


def _actual_port(endpoint: dict[str, Any] | None) -> Any:
    endpoint = endpoint or {}
    return endpoint.get("display_actual_port") or endpoint.get("actual_port")


def _location(endpoint: dict[str, Any] | None) -> Any:
    endpoint = endpoint or {}
    rack = endpoint.get("rack")
    unit = endpoint.get("unit")
    if rack and unit is not None and unit != "":
        return f"{rack}/{unit}"
    return rack or unit or NA


def _report(endpoint: dict[str, Any] | None) -> Any:
    endpoint = endpoint or {}
    return endpoint.get("report") if endpoint.get("report") is not None else endpoint.get("report_status")


def _signal_stats(endpoint: dict[str, Any] | None) -> str:
    endpoint = endpoint or {}
    if endpoint.get("is_internal"):
        return ""
    power = (endpoint.get("advanced_stats") or {}).get("power_stats")
    if not isinstance(power, dict):
        return "" if power is None else str(power)
    rx = power.get("rx_power_lane")
    tx = power.get("tx_power_lane")
    parts = []
    if rx is not None:
        rx_s = ", ".join(str(x) for x in rx) if isinstance(rx, list) else str(rx)
        parts.append(f"Rx: {rx_s}")
    if tx is not None:
        tx_s = ", ".join(str(x) for x in tx) if isinstance(tx, list) else str(tx)
        parts.append(f"Tx : {tx_s}")
    return "/".join(parts) + ("/" if parts else "")


def _ber(endpoint: dict[str, Any] | None, key: str) -> str:
    stats = ((endpoint or {}).get("advanced_stats") or {}).get("ber_stats") or {}
    value = stats.get(key)
    if value is None or value == "":
        return NA
    if isinstance(value, str):
        return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number == 0:
        return "0"
    return f"{number:.6e}"


def flatten_displayed(circuit: dict[str, Any]) -> dict[str, Any]:
    a = circuit.get("a_endpoint") or {}
    z = circuit.get("z_endpoint") or {}
    return {
        "Status": _null_as_na(circuit.get("status")),
        "Protocol": circuit.get("protocol") or "",
        "A Data Hall": _na(a.get("data_hall")),
        "A SU Number": _na(a.get("su_number")),
        "A Report": _report(a) or "",
        "A Remediation Action": a.get("remediation_action") or "",
        "A Location": _location(a),
        "Exp A Node": a.get("node") or "",
        "Exp A Port": _endpoint_port(a) or "",
        "Disc A Node": a.get("actual_node") or "",
        "Disc A Port": _actual_port(a) or "",
        "A Transceiver Connected": _null_as_na(a.get("plugged")),
        "A Port Status": _null_as_na(a.get("port_status")),
        "A Signal Stats": _signal_stats(a),
        "A Raw BER": _ber(a, "raw_ber"),
        "A Eff BER": _ber(a, "effective_ber"),
        "A Symbol BER": _ber(a, "symbol_ber"),
        "Z Data Hall": _na(z.get("data_hall")),
        "Z SU Number": _na(z.get("su_number")),
        "Z Report": _report(z) or "",
        "Z Remediation Action": z.get("remediation_action") or "",
        "Z Location": _location(z),
        "Exp Z Node": z.get("node") or "",
        "Exp Z Port": _endpoint_port(z) or "",
        "Disc Z Node": z.get("actual_node") or "",
        "Disc Z Port": _actual_port(z) or "",
        "Z Transceiver Connected": _null_as_na(z.get("plugged")),
        "Z Port Status": _null_as_na(z.get("port_status")),
        "Z Signal Stats": _signal_stats(z),
        "Z Raw BER": _ber(z, "raw_ber"),
        "Z Eff BER": _ber(z, "effective_ber"),
        "Z Symbol BER": _ber(z, "symbol_ber"),
    }


def write_displayed_csv(path: Path, circuits: Sequence[dict[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DISPLAYED_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for circuit in circuits:
            writer.writerow(flatten_displayed(circuit))
    return len(circuits)


def write_displayed_xlsx(path: Path, circuits: Sequence[dict[str, Any]]) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Circuits Issues"
    ws.append(DISPLAYED_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for circuit in circuits:
        row = flatten_displayed(circuit)
        ws.append([row.get(col, "") for col in DISPLAYED_COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DISPLAYED_COLUMNS))}{max(1, len(circuits) + 1)}"
    widths = {
        "Status": 16,
        "Protocol": 12,
        "A Report": 28,
        "Z Report": 28,
        "A Remediation Action": 40,
        "Z Remediation Action": 40,
        "Exp A Node": 24,
        "Exp Z Node": 24,
        "Disc A Node": 24,
        "Disc Z Node": 24,
        "A Signal Stats": 28,
        "Z Signal Stats": 28,
    }
    for idx, name in enumerate(DISPLAYED_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(name, 16)
    if circuits:
        table = Table(displayName="CircuitsIssues", ref=ws.auto_filter.ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    wb.save(path)
    return len(circuits)
